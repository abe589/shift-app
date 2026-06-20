import os
import io
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
import sqlite3
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'shift-app-secret-2026')

DATABASE = os.path.join(os.path.dirname(__file__), 'shifts.db')

# スタッフ一覧（Excelの列番号と対応）
STAFF_LIST = [
    '吉儀芽似',
    '新宮歩香',
    '徳田千穂',
    '加藤和瑚',
    '西尾風音',
    '嘉藤　希',
    '大國裕貴',
    '今岡和貴',
    '藤本舞',
    '岡田はな',
    '井原芽由',
    '石倉依里',
    '宮本 黎生',
]

# 各スタッフのExcel列設定（開始列, 終了列, 夜時間列）
STAFF_EXCEL_COLS = {
    '吉儀芽似':  {'start': 2,   'end': 3,   'night': 5},
    '新宮歩香':  {'start': 15,  'end': 16,  'night': 18},
    '徳田千穂':  {'start': 28,  'end': 29,  'night': 31},
    '加藤和瑚':  {'start': 41,  'end': 42,  'night': 44},
    '西尾風音':  {'start': 54,  'end': 55,  'night': 57},
    '嘉藤　希':  {'start': 67,  'end': 68,  'night': 70},
    '大國裕貴':  {'start': 80,  'end': 81,  'night': 83},
    '今岡和貴':  {'start': 93,  'end': 94,  'night': 96},
    '藤本舞':    {'start': 106, 'end': 107, 'night': 109},
    '岡田はな':  {'start': 119, 'end': 120, 'night': 122},
    '井原芽由':  {'start': 132, 'end': 133, 'night': 135},
    '石倉依里':  {'start': 145, 'end': 146, 'night': 148},
    '宮本 黎生': {'start': 158, 'end': 159, 'night': 161},
}

# 時給設定（昼/夜）
HOURLY_RATES = {
    'day': 1033,
    'night': 1291,
}


def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with get_db() as conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS shifts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_name TEXT NOT NULL,
                shift_date TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                submitted_at TEXT NOT NULL,
                UNIQUE(staff_name, shift_date)
            )
        ''')
        conn.commit()


def parse_time_to_minutes(t_str):
    """'17:00' → 1020, '25:30' → 1530"""
    h, m = map(int, t_str.split(':'))
    return h * 60 + m


def calc_night_minutes(end_str):
    """終了時刻から夜時間(分)を計算。22:00以降の部分。"""
    end_min = parse_time_to_minutes(end_str)
    night_start = 22 * 60
    if end_min <= night_start:
        return 0
    return end_min - night_start


def minutes_to_str(mins):
    """1530 → '25:30'"""
    return f"{mins // 60}:{mins % 60:02d}"


def time_to_excel_value(t_str):
    """'17:00' → Excel時間値（day分数）"""
    mins = parse_time_to_minutes(t_str)
    return mins / 60 / 24


@app.before_request
def setup():
    init_db()


@app.route('/')
def index():
    return render_template('index.html', staff_list=STAFF_LIST)


@app.route('/submit', methods=['POST'])
def submit():
    staff_name = request.form.get('staff_name', '').strip()
    shift_date = request.form.get('shift_date', '').strip()
    start_time = request.form.get('start_time', '').strip()
    end_time_raw = request.form.get('end_time', '').strip()
    is_overnight = request.form.get('is_overnight') == 'on'

    if not all([staff_name, shift_date, start_time, end_time_raw]):
        flash('すべての項目を入力してください', 'error')
        return redirect(url_for('index'))

    if staff_name not in STAFF_LIST:
        flash('スタッフ名が正しくありません', 'error')
        return redirect(url_for('index'))

    # 日跨ぎ処理：翌日0時以降の時刻は24を加算
    if is_overnight:
        h, m = map(int, end_time_raw.split(':'))
        end_time = f"{h + 24}:{m:02d}"
    else:
        end_time = end_time_raw

    # バリデーション
    try:
        start_min = parse_time_to_minutes(start_time)
        end_min = parse_time_to_minutes(end_time)
    except Exception:
        flash('時刻の形式が正しくありません (HH:MM)', 'error')
        return redirect(url_for('index'))

    if end_min <= start_min:
        flash('終了時刻は開始時刻より後にしてください', 'error')
        return redirect(url_for('index'))

    submitted_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    try:
        with get_db() as conn:
            conn.execute(
                'INSERT OR REPLACE INTO shifts (staff_name, shift_date, start_time, end_time, submitted_at) VALUES (?,?,?,?,?)',
                (staff_name, shift_date, start_time, end_time, submitted_at)
            )
            conn.commit()
        flash('入力完了しました！', 'success')
    except Exception as e:
        flash(f'エラーが発生しました: {e}', 'error')

    return redirect(url_for('confirm', staff=staff_name))


@app.route('/confirm')
def confirm():
    staff = request.args.get('staff', '')
    month = request.args.get('month', date.today().strftime('%Y-%m'))

    shifts = []
    if staff:
        with get_db() as conn:
            rows = conn.execute(
                "SELECT * FROM shifts WHERE staff_name=? AND shift_date LIKE ? ORDER BY shift_date",
                (staff, f'{month}%')
            ).fetchall()
        for r in rows:
            night_min = calc_night_minutes(r['end_time'])
            shifts.append({
                'date': r['shift_date'],
                'start': r['start_time'],
                'end': r['end_time'],
                'night': minutes_to_str(night_min) if night_min > 0 else 'なし',
            })

    return render_template('confirm.html', staff=staff, shifts=shifts, month=month)


@app.route('/admin')
def admin():
    month = request.args.get('month', date.today().strftime('%Y-%m'))

    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM shifts WHERE shift_date LIKE ? ORDER BY staff_name, shift_date",
            (f'{month}%',)
        ).fetchall()

    entries = []
    for r in rows:
        night_min = calc_night_minutes(r['end_time'])
        entries.append({
            'id': r['id'],
            'staff': r['staff_name'],
            'date': r['shift_date'],
            'start': r['start_time'],
            'end': r['end_time'],
            'night': minutes_to_str(night_min) if night_min > 0 else '0:00',
            'submitted_at': r['submitted_at'],
        })

    return render_template('admin.html', entries=entries, month=month)


@app.route('/admin/delete/<int:entry_id>', methods=['POST'])
def delete_entry(entry_id):
    month = request.form.get('month', date.today().strftime('%Y-%m'))
    with get_db() as conn:
        conn.execute('DELETE FROM shifts WHERE id=?', (entry_id,))
        conn.commit()
    return redirect(url_for('admin', month=month))


@app.route('/admin/export')
def export_excel():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    year, mon = map(int, month.split('-'))

    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM shifts WHERE shift_date LIKE ? ORDER BY shift_date",
            (f'{month}%',)
        ).fetchall()

    # ---- Excelブック生成 ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{year}年{mon}月'

    # スタイル定義
    thin = Side(style='thin', color='000000')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    subheader_fill = PatternFill('solid', fgColor='D9E1F2')
    subheader_font = Font(bold=True, size=9)
    time_fmt = '[h]:mm'

    # 月の日数
    import calendar
    days_in_month = calendar.monthrange(year, mon)[1]

    # --- ヘッダー行 ---
    # 行1: タイトル
    ws.cell(row=1, column=1, value=f'{year}年{mon}月 シフト表')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    # 行2: 列ヘッダー
    headers = ['日付', '曜日', 'スタッフ名', '開始時刻', '終了時刻', '夜時間', '勤務時間', '昼時間', '夜給与', '昼給与', '合計給与']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border_all

    WEEKDAYS = ['月', '火', '水', '木', '金', '土', '日']

    # データ行
    row_num = 3
    for r in rows:
        d = datetime.strptime(r['shift_date'], '%Y-%m-%d').date()
        weekday = WEEKDAYS[d.weekday()]
        start_min = parse_time_to_minutes(r['start_time'])
        end_min = parse_time_to_minutes(r['end_time'])
        night_min = calc_night_minutes(r['end_time'])
        work_min = end_min - start_min
        day_min = work_min - night_min

        # 日付
        ws.cell(row=row_num, column=1, value=d.strftime('%m/%d')).alignment = center
        # 曜日
        ws.cell(row=row_num, column=2, value=weekday).alignment = center
        # スタッフ名
        ws.cell(row=row_num, column=3, value=r['staff_name']).alignment = center
        # 開始時刻
        c_start = ws.cell(row=row_num, column=4)
        c_start.value = start_min / 60 / 24
        c_start.number_format = time_fmt
        c_start.alignment = center
        # 終了時刻
        c_end = ws.cell(row=row_num, column=5)
        c_end.value = end_min / 60 / 24
        c_end.number_format = time_fmt
        c_end.alignment = center
        # 夜時間
        c_night = ws.cell(row=row_num, column=6)
        c_night.value = night_min / 60 / 24 if night_min > 0 else 0
        c_night.number_format = time_fmt
        c_night.alignment = center
        # 勤務時間（計算）
        c_work = ws.cell(row=row_num, column=7)
        c_work.value = work_min / 60 / 24
        c_work.number_format = time_fmt
        c_work.alignment = center
        # 昼時間
        c_day = ws.cell(row=row_num, column=8)
        c_day.value = day_min / 60 / 24 if day_min > 0 else 0
        c_day.number_format = time_fmt
        c_day.alignment = center
        # 夜給与
        night_pay = round(night_min / 60 * HOURLY_RATES['night'])
        ws.cell(row=row_num, column=9, value=night_pay).alignment = center
        # 昼給与
        day_pay = round(day_min / 60 * HOURLY_RATES['day'])
        ws.cell(row=row_num, column=10, value=day_pay).alignment = center
        # 合計給与
        ws.cell(row=row_num, column=11, value=night_pay + day_pay).alignment = center

        # 枠線
        for col in range(1, 12):
            ws.cell(row=row_num, column=col).border = border_all

        row_num += 1

    # 列幅調整
    col_widths = [8, 5, 12, 10, 10, 8, 8, 8, 10, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- スタッフ別月次集計シート ---
    ws2 = wb.create_sheet('月次集計')
    ws2.cell(row=1, column=1, value='スタッフ名').font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value='勤務日数').font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value='合計勤務時間').font = header_font
    ws2.cell(row=1, column=3).fill = header_fill
    ws2.cell(row=1, column=4, value='合計夜時間').font = header_font
    ws2.cell(row=1, column=4).fill = header_fill
    ws2.cell(row=1, column=5, value='合計給与').font = header_font
    ws2.cell(row=1, column=5).fill = header_fill

    # スタッフ別集計
    staff_summary = {}
    with get_db() as conn:
        all_rows = conn.execute(
            "SELECT * FROM shifts WHERE shift_date LIKE ?", (f'{month}%',)
        ).fetchall()
    for r in all_rows:
        name = r['staff_name']
        start_min = parse_time_to_minutes(r['start_time'])
        end_min = parse_time_to_minutes(r['end_time'])
        night_min = calc_night_minutes(r['end_time'])
        work_min = end_min - start_min
        day_min = work_min - night_min
        pay = round(night_min / 60 * HOURLY_RATES['night']) + round(day_min / 60 * HOURLY_RATES['day'])
        if name not in staff_summary:
            staff_summary[name] = {'days': 0, 'work_min': 0, 'night_min': 0, 'pay': 0}
        staff_summary[name]['days'] += 1
        staff_summary[name]['work_min'] += work_min
        staff_summary[name]['night_min'] += night_min
        staff_summary[name]['pay'] += pay

    r2 = 2
    for name in STAFF_LIST:
        if name not in staff_summary:
            continue
        s = staff_summary[name]
        ws2.cell(row=r2, column=1, value=name)
        ws2.cell(row=r2, column=2, value=s['days'])
        c = ws2.cell(row=r2, column=3)
        c.value = s['work_min'] / 60 / 24
        c.number_format = '[h]:mm'
        c = ws2.cell(row=r2, column=4)
        c.value = s['night_min'] / 60 / 24
        c.number_format = '[h]:mm'
        ws2.cell(row=r2, column=5, value=s['pay'])
        for col in range(1, 6):
            ws2.cell(row=r2, column=col).border = border_all
            ws2.cell(row=r2, column=col).alignment = center
        r2 += 1

    # Excelをメモリに保存して送信
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'シフト_{year}年{mon}月.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/shifts')
def api_shifts():
    """スタッフが自分の入力済みシフトを確認するAPI"""
    staff = request.args.get('staff', '')
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    if not staff:
        return jsonify([])
    with get_db() as conn:
        rows = conn.execute(
            "SELECT shift_date, start_time, end_time FROM shifts WHERE staff_name=? AND shift_date LIKE ? ORDER BY shift_date",
            (staff, f'{month}%')
        ).fetchall()
    return jsonify([dict(r) for r in rows])


if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('FLASK_DEBUG', 'false').lower() == 'true')
